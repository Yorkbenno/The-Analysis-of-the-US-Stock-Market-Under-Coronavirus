import numpy as np
import pandas as pd
import openpyxl

if __name__ == '__main__':
    # This is the resource file
    resource = "../Top50 US stocks.xlsx"
    # Using 0.3 and 0.7 as two Quantile as specified in the Fama French website
    quantile1 = 0.3
    quantile2 = 0.7
    stock_num = 50
    quantile_index1 = int(stock_num * quantile1)
    quantile_index2 = int(stock_num * quantile2)
    # parse the excel file
    df = pd.read_excel(resource, sheet_name=0, header=1, skip_rows=1)
    # Do the grouping for each single factor
    factors = [["size", 2], ["B/M", 6], ["inv", 7], ["OP", 8]]
    d = {}
    for factor in factors:
        name, i = factor
        group = df.iloc[:, [0, i]]
        group = group.to_numpy(copy=True)
        if name == "B/M":
            group[:, 1] = 1.0 / group[:, 1]
            # print(group)
        group = list(group)
        group.sort(key=lambda x: x[1])
        small_group = np.array(group[:quantile_index1])[:, 0]
        middle_group = np.array(group[quantile_index1:quantile_index2])[:, 0]
        large_group = np.array(group[quantile_index2:])[:, 0]
        d[name] = np.array([small_group, middle_group, large_group])

    # print(d.items())
    # finish grouping
    # continue with two factor grouping
    keys = list(d.keys())
    two_factor_dic = {}
    for i in range(4):
        for j in range(i + 1, 4):
            # print("grouping by " + keys[i] + " and " + keys[j])
            group1 = d[keys[i]]
            group2 = d[keys[j]]
            result = []
            for k in group1:
                # result.append([])
                temp = []
                for m in group2:
                    elements = [e for e in k if e in m]
                    temp.append(elements)
                # print(temp)
                result.append(temp)

            # print(result)
            two_factor_dic[keys[i] + " & " + keys[j]] = result

    # print(two_factor_dic.items())
    # checking for correctness
    # s = set()
    # two_factor_keys = list(two_factor_dic.keys())
    # for m in two_factor_dic.values():
    #     print("===============")
    #     for rows in m:
    #         for cols in rows:
    #             print(cols, end=" ")
    #         print()
    # print(len(s))
    # print(two_factor_dic["size & OP"])
    # final = pd.DataFrame(two_factor_dic)
    # final.to_excel("./result.xlsx")

    # Do the weighting for every portfolio
    # Get the market value of each stock
    market_value_list = df.iloc[:, [0, 2]].to_numpy(copy=True)
    # create dictionary
    market_value_dict = {k: v for k, v in market_value_list}
    two_factor_keys = list(two_factor_dic.keys())
    for key in two_factor_keys:
        print("====================" + key + "====================")
        group_matrix = two_factor_dic[key]
        for rows in group_matrix:
            for cols in rows:
                market_value = []
                total = 0.0
                # Now get the list of stock of one portfolio
                for company in cols:
                    dollar = market_value_dict[company]
                    market_value.append(dollar)
                    total += dollar
                # Got the market value data
                if total == 0:
                    market_value_weights = []
                else:
                    market_value_weights = [round((m / total), 5) for m in market_value]
                # Now we have array of weights of stocks, append it
                for index, company in enumerate(cols):
                    t = tuple((company, market_value_weights[index]))
                    cols[index] = t
                print(cols, end=" ")
            print()

    # save the dictionary
    np.save("./two_factor_dic.npy", two_factor_dic)
    # write to excel
    path = "./grouping_result_final.xlsx"
    # workbook = openpyxl.load_workbook(path)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    # 6 groups
    write_row = 2
    write_col = 3
    for key in two_factor_keys:
        cell_ref = worksheet.cell(row=write_row, column=write_col)
        cell_ref.value = key
        group_matrix = two_factor_dic[key]
        for i, rows in enumerate(group_matrix):
            for j, cols in enumerate(rows):
                cell_ref = worksheet.cell(row=write_row + i + 1, column=write_col + j)
                cell_ref.value = str(cols)

        write_row += 6
    workbook.save(path)